#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
灵巧手自动化测试工具主界面
基于 PyQt5 开发，实现端口扫描、设备测试、报告导出等功能
"""
import subprocess
import sys
import os
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QVBoxLayout, QCheckBox,
    QDialog, QLabel, QStyledItemDelegate, QTextEdit, QDialogButtonBox
)
from PyQt5.QtWidgets import QMenu, QAction
from PyQt5.QtCore import Qt, QPoint, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.uic import loadUi

from rohand.rohand_logger import RoHandLogger
from rohand.rohand_manager import RohanManager
from rohand_theme import apply_black_style, apply_green_style, apply_default_style

# ---------------------------------------------------------------------------
# 公共模块导入：常量、工具函数、共享数据操作
# ---------------------------------------------------------------------------
from rohand_common import (
    COL_PORT, COL_SOFTWARE_VERSION, COL_DEVICE_ID, COL_CONNECT_STATUS, COL_TEST_RESULT,
    TABLE_HEADERS, RESULT_PASS, RESULT_FAIL, RESULT_SKIP,
    DEFAULT_TEST_RESULT,
    STATUS_CONNECTED_UI, STATUS_WAIT_TEST, STATUS_TESTING, STATUS_PAUSED,
    STATUS_UNKNOWN_DEVICE, STATUS_READ_FAIL, STATUS_NOT_CONNECTED,
    MSG_HINT, MSG_REFRESH_PORT, MSG_REFRESHING_PORTS, BTN_REFRESH, BTN_LATER, LABEL_SELECT_ALL,
    DIALOG_REFRESH_TITLE, DIALOG_REFRESH_TIP, DEFAULT_AGING_OPTIONS, REFRESH_PROMPT_DELAY_MS
)
from rohand_common import OperateSharedData


class RoHandTestWindow(QMainWindow):
    """主窗口类：集成UI显示、设备管理、测试控制、报告导出"""

    def __init__(self, parent=None):
        super().__init__(parent)

        # ===================== 全局变量初始化 =====================
        self.last_port_refresh_time = 0
        self.is_refresh_port = False
        self.protocol_type = 0

        # 端口与设备信息
        self.port_names_all = ['无可用端口']
        self.device_info_all = []
        self.port_names_selected = []
        self.selected_device_ids = []
        self.device_infos_test = []

        # 测试参数
        self.selected_aging_hours = 0.001
        self.unit_duration = 6.16
        self.total_test_seconds = 0

        # UI组件
        self.test_data_table = None
        self.check_box_list = []
        self.script_loaded = False
        self.script_path = None

        # 业务对象
        self.rohand_manager = None
        self.roHandLogger = None
        self.client = None

        # 脚本与报告
        self.script_name = None
        self.report_title = None
        self.raw_test_data = None

        # 测试状态
        self.stop_test = False
        self.pause_test = False

        # 工作线程
        self.executeScriptWorker = None
        self.deviceInfosWorker = None
        self.progressbar_worker = None

        # ===================== 加载UI文件 =====================
        base = os.path.dirname(os.path.abspath(__file__))
        ui_path = os.path.join(base, "rohand_test.ui")
        if not os.path.exists(ui_path):
            ui_path = os.path.join(base, "ui", "rohand_test.ui")
        loadUi(ui_path, self)

        # ===================== 窗口属性设置 =====================
        flags = self.windowFlags()
        flags &= ~Qt.WindowMaximizeButtonHint
        flags |= Qt.WindowSystemMenuHint | Qt.WindowTitleHint
        self.setWindowFlags(flags)

        # ===================== 初始化与绑定 =====================
        self._init_ui_widgets()
        self._bind_all_events()
        self._init_manager()

        OperateSharedData.write(False, False)
        self.closeEvent = self.close_event_handler

    def close_event_handler(self, event):
        """窗口关闭事件：清理共享数据文件"""
        OperateSharedData.delete_shared_data_file()
        event.accept()

    # ------------------------------------------------------------------
    # UI 组件初始化
    # ------------------------------------------------------------------
    def setMenuItemStyle(self, menu: QMenu):
        """设置菜单样式：统一高度、宽度、字体大小"""
        menu.setStyleSheet("""
            QMenu::item {
                height: 20px;
                width: 200px;
                margin: 0px 0px;
                font-size: 16px;
            }
        """)

    class CenterAlignDelegate(QStyledItemDelegate):
        """表格内容居中代理：修复PyQt5对齐不生效问题"""
        def initStyleOption(self, option, index):
            super().initStyleOption(option, index)
            option.displayAlignment = Qt.AlignCenter

    def _init_ui_widgets(self):
        """初始化所有UI组件样式与默认值"""
        self.setMenuItemStyle(self.menu_file)
        self.setMenuItemStyle(self.menu_config)
        self.setMenuItemStyle(self.menu_theme)
        self.setMenuItemStyle(self.menu_help)

        # 日志框
        self.log_text_edit.setReadOnly(True)
        self.log_text_edit.setFont(QFont("Consolas", 10))

        # 测试数据表格
        self.test_data_table = QTableWidget(self.test_data_group)
        self.test_data_table.setGeometry(10, 30, 800, 580)
        self.test_data_table.setColumnCount(5)
        self.test_data_table.setHorizontalHeaderLabels(TABLE_HEADERS)
        self.test_data_table.setRowCount(0)
        self.test_data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.test_data_table.setItemDelegate(self.CenterAlignDelegate())

        # 老化时间下拉框
        self.aging_time_combo.clear()
        self.aging_time_combo.addItems(DEFAULT_AGING_OPTIONS)

        # 进度条
        self.test_progress_bar.setRange(0, 100)
        self.test_progress_bar.setValue(0)
        self.test_progress_bar.setStyleSheet("QProgressBar { color: #000000; font-weight: bold; }")

        # 统计数据初始化
        self.total_case_value.setText("0条")
        self.success_case_value.setText("0条")
        self.fail_case_value.setText("0条")
        self.skip_case_value.setText("0条")

        # 状态栏
        self.status_bar.showMessage(MSG_REFRESH_PORT)

    # ------------------------------------------------------------------
    # 组件事件绑定
    # ------------------------------------------------------------------
    def _bind_all_events(self):
        """绑定所有按钮、菜单、组件的信号与槽函数"""
        self.aging_time_combo.currentTextChanged.connect(self.on_aging_time_selected)
        self.log_copy_btn.clicked.connect(self.on_copy_log)
        self.log_clear_btn.clicked.connect(self.on_clear_log)
        self.log_save_btn.clicked.connect(self.on_save_log)
        self.refresh_btn.clicked.connect(self.start_port_refresh)
        self.select_all_check.stateChanged.connect(self.on_select_all)
        self.start_test_btn.clicked.connect(self.on_start_test)
        self.pause_test_btn.clicked.connect(self.on_pause_test)
        self.stop_test_btn.clicked.connect(self.on_stop_test)
        self.log_level_btn.clicked.connect(self.on_log_level_clicked)

        # 文件菜单
        self.action_loadScript.triggered.connect(self.on_load_script)
        self.action_exportReport.triggered.connect(self.on_export_report)
        self.action_exit.triggered.connect(self.close)

        # 配置菜单
        self.action_viewConfigFile.triggered.connect(self.on_view_config)
        self.action_editConfigFile.triggered.connect(self.on_edit_config)

        # 主题菜单
        self.action_blackTheme.triggered.connect(self.on_theme_black)
        self.action_greenTheme.triggered.connect(self.on_theme_green)
        self.action_defaultTheme.triggered.connect(self.on_theme_default)

        # 帮助菜单
        self.action_about.triggered.connect(self.on_about)

    # ------------------------------------------------------------------
    # 业务管理器初始化
    # ------------------------------------------------------------------
    def _init_manager(self):
        """初始化日志、协议、设备管理对象"""
        print('_init_manager')
        self.rologger = RoHandLogger(self.log_text_edit)
        self.protocol_type = int(RohanManager.read_config_value(
            section="protocol_type", key="protocol", default=0
        ))
        self.rohand_manager = RohanManager(self.protocol_type)

    # ------------------------------------------------------------------
    # 控件响应函数
    # ------------------------------------------------------------------
    def on_aging_time_selected(self, text):
        """老化时间选择事件"""
        try:
            hour_value = float(text.replace("小时", ""))
            self.rologger.log(f"已选择老化时间：{text} → 提取数字：{hour_value} 小时")
            self.selected_aging_hours = hour_value
        except Exception as e:
            self.rologger.log(f"选择时间异常：{e}")

    def on_copy_log(self):
        """复制日志内容到剪贴板"""
        self.rologger.log('on_copy_log')
        self.log_text_edit.selectAll()
        self.log_text_edit.copy()

    def on_clear_log(self):
        """清空日志显示框"""
        self.rologger.log('on_clear_log')
        self.log_text_edit.clear()

    def on_save_log(self):
        """保存日志到本地文件"""
        self.rologger.log("开始保存日志文件")
        filename = f"灵巧手测试日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        fn, _ = QFileDialog.getSaveFileName(self, "保存日志", filename, "文本文件 (*.txt)")

        if fn:
            try:
                with open(fn, "w", encoding="utf-8") as f:
                    f.write(self.log_text_edit.toPlainText())
                self.rologger.log(f"日志保存成功：{os.path.basename(fn)}")
            except Exception as e:
                self.rologger.log(f"日志保存失败：{str(e)}", level="ERROR")

    def on_log_level_clicked(self):
        """日志级别按钮：弹出级别选择菜单"""
        self.rologger.log(f"on_log_level_clicked")
        menu = QMenu(self)
        self.setMenuItemStyle(menu)
        log_levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']
        current_level = getattr(self.rologger, 'log_level', 'INFO')

        for level in log_levels:
            action = QAction(level, self)
            if level == current_level:
                action.setCheckable(True)
                action.setChecked(True)
            action.triggered.connect(lambda checked, l=level: self.on_log_level_selected(l))
            menu.addAction(action)

        pos = self.log_level_btn.mapToGlobal(QPoint(0, self.log_level_btn.height()))
        menu.exec_(pos)

    def on_log_level_selected(self, level):
        """设置日志输出级别"""
        try:
            self.rologger.set_log_level(level)
            self.rologger.log(f"日志级别已设置为：{level}")
        except Exception as e:
            self.rologger.log(f"设置日志级别失败：{str(e)}")

    # ------------------------------------------------------------------
    # 端口与设备管理工具函数
    # ------------------------------------------------------------------
    def remove_all_widgets_from_layout(self, layout):
        """清空并删除布局内所有控件与子布局"""
        if layout is None:
            return

        while layout.count() > 0:
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
            elif item.layout():
                self.remove_all_widgets_from_layout(item.layout())

    def update_port_info(self):
        """刷新端口复选框列表UI"""
        self.rologger.log('update_port_info')
        max_per_col, max_col = 8, 4
        vertical_layouts = []

        # 清空旧端口控件
        port_layout = self.scroll_content_widget.layout()
        self.remove_all_widgets_from_layout(port_layout)
        self.check_box_list.clear()

        # 重新生成端口复选框
        for idx, port in enumerate(self.port_names_all):
            cbx = QCheckBox(port)
            cbx.setObjectName("port_checkbox")
            self.check_box_list.append(cbx)

            col = idx // max_per_col
            if col >= len(vertical_layouts):
                vl = QVBoxLayout()
                vl.setAlignment(Qt.AlignTop)
                vl.setSpacing(8)
                port_layout.addLayout(vl)
                vertical_layouts.append(vl)

            vertical_layouts[col].addWidget(cbx)
            cbx.clicked.connect(lambda checked, cb=cbx: self.on_port_cbx_clicked(checked, cb))

        port_layout.setContentsMargins(12, 8, 18, 8)
        port_layout.setSpacing(10)

    def on_port_cbx_clicked(self, checked, checkbox):
        """单个端口复选框点击事件"""
        self.rologger.log('on_port_cbx_clicked')
        try:
            port = checkbox.text()
            if checked and port not in self.port_names_selected:
                self.port_names_selected.append(port)
            elif not checked and port in self.port_names_selected:
                self.port_names_selected.remove(port)

            self.update_test_datas_table()
            self.rologger.log(f"端口 {port} {'选中' if checked else '取消选中'}，已选：{self.port_names_selected}")
        except Exception as e:
            self.rologger.log(f"复选框操作失败：{str(e)}")

    def update_test_datas_table(self):
        """根据选中端口更新测试设备表格"""
        self.rologger.log('update_test_datas_table')
        self.device_infos_test.clear()
        self.test_data_table.setRowCount(0)

        if not self.port_names_selected or not self.device_info_all:
            return

        # 按选中端口筛选设备
        for port in self.port_names_selected:
            for dev in self.device_info_all:
                if dev.get(COL_PORT) == port:
                    self.device_infos_test.append(dev)
                    break

        self._on_device_info_update(self.device_infos_test)

    def get_row_by_port(self, port):
        """根据端口号查找表格行号，未找到返回-1"""
        for row in range(self.test_data_table.rowCount()):
            item = self.test_data_table.item(row, 0)
            if item and item.text() == port:
                return row
        return -1

    def _on_device_info_update(self, device_info_test):
        """设备信息更新：刷新表格显示"""
        self.rologger.log('_on_device_info_update')
        self.test_data_table.setRowCount(0)

        if not device_info_test:
            self.rologger.log("无设备信息需要展示")
            return

        for device_info in device_info_test:
            port = device_info.get(COL_PORT)
            if not port:
                continue

            row = self.test_data_table.rowCount()
            self.test_data_table.insertRow(row)

            self.test_data_table.setItem(row, 0, QTableWidgetItem(port))
            self.test_data_table.setItem(row, 1, QTableWidgetItem(device_info.get(COL_SOFTWARE_VERSION, "")))
            self.test_data_table.setItem(row, 2, QTableWidgetItem(str(device_info.get(COL_DEVICE_ID, ""))))
            self.test_data_table.setItem(row, 3, QTableWidgetItem(device_info.get(COL_CONNECT_STATUS, "")))
            self.test_data_table.setItem(row, 4, QTableWidgetItem(device_info.get(COL_TEST_RESULT, "待测试")))

            self.rologger.log(f'已添加设备到表格：{port}')

    def update_table_column_by_port(self, port, column_index, value):
        """根据端口号更新表格指定列的值"""
        row = self.get_row_by_port(port)
        if row >= 0:
            self.test_data_table.setItem(row, column_index, QTableWidgetItem(str(value)))
            self.rologger.log(f"更新表格 → 端口 {port} 第 {column_index} 列为：{value}")
        else:
            self.rologger.log(f"更新失败：端口 {port} 不存在于表格中")

    # ------------------------------------------------------------------
    # 端口刷新
    # ------------------------------------------------------------------
    def _on_port_refresh_finished(self, device_infos):
        """端口刷新线程完成回调"""
        self.rologger.log(f'_on_port_refresh_finished, device_infos = {device_infos}')

        # 清空旧数据
        self.port_names_all.clear()
        self.port_names_selected.clear()
        self.device_info_all.clear()
        self.device_infos_test.clear()
        self.test_data_table.setRowCount(0)

        # 提取端口列表
        if device_infos:
            self.port_names_all = [dev.get(COL_PORT, "") for dev in device_infos if dev.get(COL_PORT)]
        else:
            self.port_names_all = ["无可用端口"]

        self.device_info_all = device_infos
        self.rologger.log(f'端口列表：{self.port_names_all}')

        if self.port_names_all and self.port_names_all[0] != "无可用端口":
            self.status_bar.showMessage("端口刷新完成")
            self.update_port_info()
        else:
            self.status_bar.showMessage("无可用端口")

        self.set_controls_enabled(True)
        self.is_refresh_port = False

    def set_controls_enabled(self, enabled):
        """批量设置控制按钮启用/禁用状态"""
        controls = [
            self.refresh_btn,
            self.aging_time_combo,
            self.select_all_check,
            self.start_test_btn,
            self.pause_test_btn,
            self.stop_test_btn
        ]
        for control in controls:
            control.setEnabled(enabled)

    def start_port_refresh(self):
        """启动端口刷新（5秒冷却+防重复刷新）"""
        self.rologger.log(f'start_port_refresh')
        current_time = time.time()

        # 5秒内限制刷新
        if current_time - self.last_port_refresh_time < 5:
            self.rologger.log(f"端口刷新过于频繁，请等待5秒后再试")
            self.status_bar.showMessage(f"操作频繁，请5秒后再刷新端口")
            return

        # 正在刷新中，直接返回
        if self.is_refresh_port:
            return

        # 记录刷新时间
        self.last_port_refresh_time = current_time
        self.is_refresh_port = True
        self.set_controls_enabled(False)

        # 清空布局
        port_layout = self.scroll_content_widget.layout()
        self.remove_all_widgets_from_layout(port_layout)
        self.check_box_list.clear()

        # 清空所有数据
        self.port_names_all.clear()
        self.port_names_selected.clear()
        self.device_info_all.clear()
        self.device_infos_test.clear()
        self.test_data_table.setRowCount(0)

        # 启动刷新线程
        self.deviceInfosWorker = DeviceInfosWorker()
        self.deviceInfosWorker.finished_with_device_infos.connect(self._on_port_refresh_finished)
        self.deviceInfosWorker.start()
        self.status_bar.showMessage(f"端口刷新中，请耐心等待...")

    # ------------------------------------------------------------------
    # 全选/取消全选
    # ------------------------------------------------------------------
    def on_select_all(self, state):
        """全选复选框状态改变事件"""
        self.rologger.log(f'on_select_all')
        checked = (state == 2)
        if self.port_names_all[0] == '无可用端口':
            return

        # 取消选中时清空表格
        if not checked:
            self.test_data_table.setRowCount(0)
            self.port_names_selected.clear()

        # 批量设置复选框状态
        for cbx in self.check_box_list:
            cbx.blockSignals(True)
            cbx.setChecked(checked)
            cbx.blockSignals(False)

        # 更新选中列表
        self.port_names_selected = self.port_names_all.copy() if checked else []
        self.update_test_datas_table()

    # ------------------------------------------------------------------
    # 设备ID与进度条偏差
    # ------------------------------------------------------------------
    def get_device_ids(self, device_infos_test):
        """从设备信息中提取设备ID列表"""
        device_ids = []
        if not device_infos_test:
            return device_ids

        for info in device_infos_test:
            dev_id = info.get(COL_DEVICE_ID)
            if dev_id is not None and dev_id != "":
                device_ids.append(int(dev_id))

        return device_ids

    def get_offset_duration(self):
        """计算补时时长，使总时长为完整周期"""
        self.rologger.log(f'get_offset_duration')
        value = float(RohanManager.read_config_value(
            section="aging_parameter", key="unit_duration", default=0
        ))
        self.unit_duration = float(str(value).strip())

        if self.unit_duration <= 0:
            return 0.0

        total_seconds = float(self.selected_aging_hours) * 3600
        total_cycles = total_seconds / self.unit_duration
        decimal_part = total_cycles - int(total_cycles)
        offset_seconds = round((1 - decimal_part) * self.unit_duration, 2)

        self.rologger.log(f'offset_seconds = {offset_seconds}')
        return offset_seconds

    # ------------------------------------------------------------------
    # 测试控制：开始 / 暂停 / 停止
    # ------------------------------------------------------------------
    def on_start_test(self):
        """开始测试按钮点击事件"""
        self.rologger.log('on_start_test')
        self.rologger.log(f"script_name = {self.script_name}")

        if not self.port_names_selected:
            self.status_bar.showMessage(f"请选择要测试的端口")
            return

        if not self.script_name:
            self.status_bar.showMessage(f"请选择测试脚本")
            return

        if self.executeScriptWorker is not None and self.executeScriptWorker.isRunning():
            return

        # 重置测试状态
        self.pause_test = False
        self.stop_test = False
        OperateSharedData.write(False, False)

        # 计算总测试时长
        self.total_test_seconds = self.selected_aging_hours * 3600 + self.get_offset_duration()

        # 获取设备ID
        self.selected_device_ids = self.get_device_ids(self.device_infos_test)

        # 启动脚本执行线程
        self.executeScriptWorker = ExecuteScriptWorker(
            self.port_names_selected,
            self.selected_device_ids,
            self.selected_aging_hours,
            self.script_name
        )
        self.executeScriptWorker.finished_with_script_result.connect(self._on_script_finished_result)
        self.executeScriptWorker.start()

        # 启动进度条线程
        self.test_progress_bar.setRange(0, 100)
        self.test_progress_bar.setValue(0)
        self.progressbar_worker = ProgressBarWorker(duration=self.total_test_seconds)
        self.progressbar_worker.progress_update.connect(self._on_progress_result)
        self.progressbar_worker.finished_signal.connect(self.on_test_finished_auto)
        self.progressbar_worker.start()

        self.status_bar.showMessage("开始测试...")

    def _on_progress_result(self, value):
        """更新进度条显示"""
        self.test_progress_bar.setValue(value)

    def on_test_finished_auto(self):
        """测试时间结束自动完成"""
        self.status_bar.showMessage("测试已完成")
        self.rologger.log("老化测试时间到，自动结束")

    def _on_script_finished_result(self, titles, result):
        """脚本执行完成，解析并更新测试结果"""
        self.rologger.log(f'_on_script_finished_result')
        self.rologger.log(f'titles = {titles},result = {result}')

        self.report_title = titles
        self.raw_test_data = result

        # 按端口分组解析结果
        port_data_dict = self.parse_test_result(result)

        # 更新每个端口的最终测试状态
        for port, table_rows in port_data_dict.items():
            final_result = "通过"
            for row_data in table_rows:
                if row_data[2] != "通过":
                    final_result = "不通过"
                    break
            self.update_table_column_by_port(port, 4, final_result)
            self.rologger.log(f"端口 {port} 最终测试结果：{final_result}")

        self._update_test_result(self.raw_test_data)

    def _update_test_result(self, raw_test_data):
        """更新测试统计数据与状态标签"""
        self.rologger.log(f'_update_test_result')
        port_data_dict = self.parse_test_result(raw_test_data)

        total_count = 0
        success_count = 0
        failed_count = 0

        for port, table_rows in port_data_dict.items():
            for row_data in table_rows:
                test_result = row_data[2]
                total_count += 1
                if test_result == "通过":
                    success_count += 1
                else:
                    failed_count += 1

        # 更新UI统计
        self.rologger.log(f'total_count={total_count},success_count={success_count},failed_count={failed_count}')
        self.total_case_value.setText(str(total_count))
        self.success_case_value.setText(str(success_count))
        self.fail_case_value.setText(str(failed_count))

        # 更新状态标签
        if failed_count == 0 and total_count > 0:
            self.test_status_label.setStyleSheet("color: #2E8B57; font-size:14px; font-weight:bold;")
            self.test_status_label.setText("✅ 测试完成 | 全部用例通过")
        elif total_count == 0:
            self.test_status_label.setStyleSheet("color: #696969; font-size:14px;")
            self.test_status_label.setText("⌛ 等待测试 | 暂无用例执行")
        else:
            self.test_status_label.setStyleSheet("color: #DC143C; font-size:14px; font-weight:bold;")
            self.test_status_label.setText("❌ 测试完成 | 存在失败用例")

    def parse_test_result(self, result):
        """解析测试结果，按端口分组返回表格数据"""
        self.rologger.log('开始解析测试结果数据，用于表格刷新')
        port_data_dict = {}

        for item in result:
            port = item['port']
            if port not in port_data_dict:
                port_data_dict[port] = []

            for gesture in item['gestures']:
                table_row = (
                    gesture['timestamp'],
                    gesture['content'],
                    gesture['result'],
                    gesture['comment']
                )
                port_data_dict[port].append(table_row)

        self.rologger.log(f'测试结果解析完成，共解析到端口：{list(port_data_dict.keys())}')
        return port_data_dict

    def on_pause_test(self):
        """测试暂停/恢复切换"""
        self.rologger.log('on_pause_test')

        if self.executeScriptWorker is None:
            self.status_bar.showMessage('当前没有在执行的任务')
            return

        if self.stop_test:
            self.status_bar.showMessage('当前无测试脚本在运行')
            return

        self.pause_test = not self.pause_test
        OperateSharedData.write(stop_test=False, pause_test=self.pause_test)

        if self.pause_test:
            self.progressbar_worker.pause()
            self.pause_test_btn.setText("继续测试")
            self.status_bar.showMessage('当前任务已暂停，点击继续执行')
            self.rologger.log("测试已暂停")
        else:
            self.progressbar_worker.resume()
            self.pause_test_btn.setText("暂停测试")
            self.status_bar.showMessage('当前任务已恢复执行')
            self.rologger.log("测试已恢复")

    def on_stop_test(self):
        """停止当前测试任务"""
        self.rologger.log(f'on_stop_test')

        if self.executeScriptWorker is None:
            self.status_bar.showMessage('当前没有在执行的任务')
            return

        self.stop_test = True
        self.pause_test = False
        OperateSharedData.write(stop_test=True, pause_test=False)

        self.status_bar.showMessage("正在停止任务...")
        self.progressbar_worker.stop()

        def _on_task_finished(title, result):
            self.on_test_finished_auto()
            self.status_bar.showMessage("当前任务已停止")
            self.stop_test = False
            self.pause_test = False
            self.executeScriptWorker.finished_with_script_result.disconnect(_on_task_finished)

        self.executeScriptWorker.finished_with_script_result.connect(_on_task_finished)
        self.executeScriptWorker.stop_task()

    # ------------------------------------------------------------------
    # 加载脚本与输出报告
    # ------------------------------------------------------------------
    def on_load_script(self):
        """加载测试脚本"""
        if self.pause_test:
            self.status_bar.showMessage('当前有任务正在在运行，请勿重复加载脚本')
            return
        self.rologger.log(f'on_load_script')
        base_dir = os.path.dirname(os.path.abspath(__file__))
        scripts_dir = os.path.join(base_dir, "scripts")

        if not os.path.exists(scripts_dir):
            os.makedirs(scripts_dir)

        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择要执行的脚本', scripts_dir, 'Python files (*.py)'
        )

        if file_path:
            self.script_name = file_path
            self.rologger.log(f'成功加载脚本：{self.script_name}')
            self.rologger.log(f'请点击【开始测试】按钮执行脚本\n')

    def on_export_report(self):
        """导出测试报告为Excel文件"""
        self.rologger.log(f'on_export_report')
        if not self.raw_test_data:
            self.rologger.log("导出失败：无测试数据")
            self.status_bar.showMessage(f"尚未进行测试，无数据可导出")
            return

        if not self.report_title:
            self.rologger.log("导出失败：报告标题为空")
            self.status_bar.showMessage(f"请设置报告标题")
            return

        # 表格样式定义
        headers = ["用例编号", "时间戳", "内容", "测试结果", "备注", "端口号"]
        column_widths = [10, 25, 35, 12, 15, 18]
        default_row_height = 35
        thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "测试报告"

        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = self.report_title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = alignment
        ws.row_dimensions[1].height = default_row_height

        # 基础信息行
        ws['A2'] = '产品名称'
        ws['C2'] = '产品型号'
        ws.merge_cells('D2:F2')
        ws['A3'] = '测试地点'
        ws['C3'] = '测试时间'
        ws.merge_cells('D3:F3')
        ws['A4'] = '测试设备'
        ws['C4'] = '设备编号'
        ws.merge_cells('D4:F4')
        ws['A5'] = '温度'
        ws['C5'] = '湿度'
        ws.merge_cells('D5:F5')

        # 表头
        ws.append(headers)
        ws.row_dimensions[6].height = default_row_height

        # 写入数据
        row_index = 7
        for item in self.raw_test_data:
            try:
                port = item["port"]
                for g in item["gestures"]:
                    ts = g.get("timestamp", "")
                    content = str(g.get("content", ""))
                    res = g.get("result", "")
                    comment = g.get("comment", "")

                    ws.append([row_index - 6, ts, content, res, comment, port])
                    row_index += 1
            except Exception as e:
                self.rologger.log(f"数据处理异常: {e}")

        # 设置列宽与样式
        for i, w in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.border = thin_border

        # 保存文件
        os.makedirs("report", exist_ok=True)
        name = os.path.splitext(os.path.basename(self.script_name))[0]
        filename = f"{name}_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join("report", filename)

        wb.save(filepath)
        QMessageBox.information(self, "成功", f"报告已保存：\n{filepath}")

    # ------------------------------------------------------------------
    # 配置文件操作
    # ------------------------------------------------------------------
    def on_view_config(self):
        """查看配置文件（只读）"""
        self.rologger.log(f'on_view_config')
        config_path = RohanManager.get_configfile_path()

        if not os.path.exists(config_path):
            QMessageBox.warning(self, "警告", f"配置文件不存在：{config_path}")
            self.rologger.log(f"配置文件不存在：{config_path}")
            return

        try:
            with open(config_path, 'r', encoding='UTF-8') as f:
                content = f.read()

            dialog = QDialog(self)
            dialog.setWindowTitle("查看配置文件")
            dialog.setMinimumSize(600, 800)
            layout = QVBoxLayout()

            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setFont(QFont("Consolas", 10))
            text_edit.setPlainText(content)
            text_edit.setStyleSheet("""
                QTextEdit {
                    background-color: white;
                    color: #333333;
                    border: 2px solid #d1d5db;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                }
            """)

            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(text_edit)
            layout.addWidget(button_box)
            dialog.setLayout(layout)
            dialog.exec_()

            self.rologger.log(f"成功查看配置文件：{config_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取配置文件失败：{str(e)}")
            self.rologger.log(f"读取配置文件失败：{str(e)}")

    def on_edit_config(self):
        """编辑配置文件（可写）"""
        self.rologger.log(f'on_edit_config')
        config_path = RohanManager.get_configfile_path()

        if not os.path.exists(config_path):
            QMessageBox.warning(self, "警告", f"配置文件不存在：{config_path}")
            self.rologger.log(f"配置文件不存在：{config_path}")
            return

        try:
            with open(config_path, 'r', encoding='UTF-8') as f:
                content = f.read()

            dialog = QDialog(self)
            dialog.setWindowTitle("修改配置文件")
            dialog.setMinimumSize(600, 800)
            layout = QVBoxLayout()

            hint_label = QLabel("提示：修改后需要重启应用程序才能生效")
            hint_label.setStyleSheet("color: #f59e0b; font-weight: bold; padding: 5px;")
            layout.addWidget(hint_label)

            text_edit = QTextEdit()
            text_edit.setFont(QFont("Consolas", 10))
            text_edit.setPlainText(content)
            text_edit.setStyleSheet("""
                QTextEdit {
                    background-color: white;
                    color: #333333;
                    border: 2px solid #d1d5db;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                }
            """)

            button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
            save_btn = button_box.button(QDialogButtonBox.Save)
            save_btn.setText("保存")
            save_btn.clicked.connect(lambda: self._save_config_and_restart(config_path, text_edit.toPlainText(), dialog))

            cancel_btn = button_box.button(QDialogButtonBox.Cancel)
            cancel_btn.setText("取消")
            cancel_btn.clicked.connect(dialog.reject)

            layout.addWidget(text_edit)
            layout.addWidget(button_box)
            dialog.setLayout(layout)
            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"编辑配置文件失败：{str(e)}")
            self.rologger.log(f"编辑配置文件失败：{str(e)}")

    def _save_config_and_restart(self, config_path, content, dialog):
        """保存配置并重启程序"""
        try:
            with open(config_path, 'w', encoding='UTF-8') as f:
                f.write(content)

            self.rologger.log(f"配置文件已保存：{config_path}")
            QMessageBox.information(self, "成功", "配置文件已保存，应用程序将自动重启...")
            dialog.accept()

            script_path = os.path.abspath(sys.argv[0])
            subprocess.Popen(
                [sys.executable, script_path],
                start_new_session=True,
                creationflags=subprocess.DETACHED_PROCESS if os.name == 'nt' else 0
            )

            time.sleep(0.5)
            QApplication.quit()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")
            self.rologger.log(f"保存配置文件失败：{str(e)}")

    # ------------------------------------------------------------------
    # 主题切换
    # ------------------------------------------------------------------
    def on_theme_black(self):
        self.rologger.log("on_theme_black")
        apply_black_style(self)

    def on_theme_green(self):
        self.rologger.log("on_theme_green")
        apply_green_style(self)

    def on_theme_default(self):
        self.rologger.log(f"on_theme_default")
        apply_default_style(self)

    # ------------------------------------------------------------------
    # 关于
    # ------------------------------------------------------------------
    def on_about(self):
        QMessageBox.about(
            self,
            "关于",
            "灵巧手自动化测试工具 v1.0\n\n"
            "基于 PyQt5 + pytest 开发\n"
            "© 2026 上海傲意科技 All Rights Reserved."
        )


# ---------------------------------------------------------------------------
# 工作线程类
# ---------------------------------------------------------------------------
class DeviceInfosWorker(QThread):
    """设备信息扫描线程：获取端口 + 设备信息"""
    finished_with_device_infos = pyqtSignal(list)

    def run(self):
        try:
            protocol_type = int(RohanManager.read_config_value(section="protocol_type", key="protocol", default=0))
            manager = RohanManager(protocol_type)
            ports = manager.read_port_info()
            device_infos = manager.get_device_info_list(ports)
            manager.delete_client()
            self.finished_with_device_infos.emit(device_infos)
        except Exception as e:
            print(f'DeviceInfosWorker ,发生异常{str(e)}')
            self.finished_with_device_infos.emit([])
            self.finished_with_device_infos.emit([])


class ExecuteScriptWorker(QThread):
    """测试脚本执行线程"""
    finished_with_script_result = pyqtSignal(str, list)

    def __init__(self, ports: list, device_ids: list, aging_duration: float, script_path: str, parent=None):
        super().__init__(parent)
        self.ports = ports
        self.device_ids = device_ids
        self.aging_duration = aging_duration
        self.script_path = script_path
        self._is_stopped = False

    def run(self):
        try:
            if self._is_stopped:
                self.finished_with_script_result.emit("任务已停止", [])
                return

            if not os.path.exists(self.script_path):
                self.finished_with_script_result.emit("脚本不存在", [])
                return

            script_dir = os.path.dirname(self.script_path)
            if script_dir not in sys.path:
                sys.path.append(script_dir)

            script_name = os.path.splitext(os.path.basename(self.script_path))[0]
            module = __import__(script_name)

            report_title, overall_result = module.main(
                ports=self.ports,
                devices_ids=self.device_ids,
                aging_duration=self.aging_duration
            )

            if self._is_stopped:
                self.finished_with_script_result.emit("任务已停止", [])
                return

            self.finished_with_script_result.emit(report_title, overall_result)

        except Exception as e:
            print(f'ExecuteScriptWorker发生异常{str(e)}')
            if self._is_stopped:
                self.finished_with_script_result.emit("任务已停止", [])
            else:
                self.finished_with_script_result.emit(f"执行异常：{str(e)}", [])

    def stop_task(self):
        self._is_stopped = True


class ProgressBarWorker(QThread):
    """进度条计时线程"""
    progress_update = pyqtSignal(int)
    finished_signal = pyqtSignal()

    def __init__(self, duration, parent=None):
        super().__init__(parent)
        self.total_test_seconds = max(duration, 1)
        self.is_running = True
        self.is_paused = False
        self.delay = 0.0

    def run(self):
        start_time = time.time()
        while self.is_running:
            if self.is_paused:
                time.sleep(0.1)
                self.delay += 0.1
                continue

            elapsed = time.time() - start_time - self.delay
            progress = int((elapsed / self.total_test_seconds) * 100)
            progress = max(0, min(100, progress))

            self.progress_update.emit(progress)
            if progress >= 100:
                break

            time.sleep(0.5)

        self.progress_update.emit(100)
        self.finished_signal.emit()

    def pause(self):
        self.is_paused = True

    def resume(self):
        self.is_paused = False

    def stop(self):
        self.is_running = False


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = RoHandTestWindow()
    win.setWindowTitle("灵巧手自动化测试界面")
    win.show()
    sys.exit(app.exec_())