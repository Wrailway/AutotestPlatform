#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
灵巧手自动化测试工具主界面
基于 PyQt5 开发，实现端口扫描、设备测试、报告导出等功能
"""
import subprocess
import sys
import os
import time
from datetime import datetime

import pytest
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.fonts import Font
from openpyxl import Workbook  # 补上缺失的导入

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QVBoxLayout,
    QDialog, QLabel, QStyledItemDelegate, QTextEdit, QDialogButtonBox
)
from PyQt5.QtWidgets import QMenu, QAction
from PyQt5.QtCore import Qt, QPoint, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QBrush, QColor
from PyQt5.uic import loadUi

from server.server_common import TABLE_HEADERS, DEFAULT_EXECUTE_TIMES_OPTIONS, DEFAULT_OPERATE_INTERVAL_OPTIONS, \
    OperateSharedData, DEFAULT_OPERATE_ENVIRONMENT_OPTIONS, DEFAULT_NUM_OF_THREADS_OPTIONS, \
    DEFAULT_NUM_OF_CONCURRENT_USERS_OPTIONS, DEFAULT_DURATION_OPTIONS, DEFAULT_RAMP_UP_OPTIONS, ENVIRONMENT_TEST
from server.server_logger import ServerHandLogger
from server.server_manager import ServerManager
from server.server_theme import apply_black_style, apply_green_style, apply_default_style


# ---------------------------------------------------------------------------
# ====================== Pytest 插件 & 执行线程 ======================
# ---------------------------------------------------------------------------
class PytestCollectPlugin:
    def __init__(self):
        self.collected_items = []

    def pytest_collection_modifyitems(self, session, config, items):
        for idx, item in enumerate(items):
            doc = item.obj.__doc__.strip() if item.obj.__doc__ else item.name
            self.collected_items.append({
                "index": idx,
                "nodeid": item.nodeid,
                "name": doc
            })


class TestCollectorThread(QThread):
    sig_collected = pyqtSignal(list)
    sig_error = pyqtSignal(str)

    def __init__(self, script_path, parent=None):
        super().__init__(parent)
        self.script_path = script_path

    def run(self):
        try:
            plugin = PytestCollectPlugin()
            pytest.main([
                "--collect-only",
                "-q",
                "-p", "no:logging",
                "-p", "no:faulthandler",
                self.script_path
            ], plugins=[plugin])
            self.sig_collected.emit(plugin.collected_items)
        except Exception as e:
            self.sig_error.emit(str(e))


class TestRunnerThread(QThread):
    sig_log = pyqtSignal(dict)
    sig_status = pyqtSignal(int, str)
    sig_finished = pyqtSignal()

    def __init__(self, script_path, node_to_index, parent=None):
        super().__init__(parent)
        self.script_path = script_path
        self.node_to_index = node_to_index
        self._pause = False
        self._stop = False

    def run(self):
        try:
            plugin = PytestRunnerPlugin(
                self.sig_log,
                self.sig_status,
                self.node_to_index,
                self
            )
            pytest.main([
                "-q", "-s", "--tb=short",
                "-p", "no:logging",
                "-p", "no:faulthandler",
                "-p", "no:cacheprovider",
                self.script_path
            ], plugins=[plugin])
        except Exception:
            pass
        finally:
            self.sig_finished.emit()

    def set_pause(self, val: bool):
        self._pause = val

    def set_stop(self, val: bool):
        self._stop = val

    def is_paused(self):
        return self._pause

    def is_stopped(self):
        return self._stop


class PytestRunnerPlugin:
    def __init__(self, log_signal, status_signal, node_to_index, runner_thread):
        self.log_signal = log_signal
        self.status_signal = status_signal
        self.node_to_index = node_to_index
        self.runner_thread = runner_thread

    @pytest.fixture
    def logger(self):
        return lambda msg: self.log_signal.emit({"level": "info", "msg": str(msg)})

    def pytest_runtest_setup(self, item):
        while self.runner_thread.is_paused() and not self.runner_thread.is_stopped():
            time.sleep(0.05)

        if self.runner_thread.is_stopped():
            pytest.skip("手动停止")

        idx = self.node_to_index.get(item.nodeid)
        if idx is not None:
            self.log_signal.emit({"level": "INFO", "msg": f"开始执行: {item.name}"})
            self.status_signal.emit(idx, "RUNNING")

    def pytest_runtest_logreport(self, report):
        if self.runner_thread.is_stopped():
            return

        idx = self.node_to_index.get(report.nodeid)
        if idx is None or report.when != "call":
            return

        if report.passed:
            self.status_signal.emit(idx, "PASS")
            self.log_signal.emit({"level": "success", "msg": f"用例[{idx+1}]通过"})
        elif report.failed:
            self.status_signal.emit(idx, "FAIL")
            self.log_signal.emit({"level": "error", "msg": f"用例[{idx+1}]失败"})
        elif report.skipped:
            self.status_signal.emit(idx, "SKIP")
            self.log_signal.emit({"level": "skip", "msg": f"用例[{idx+1}]跳过"})


# ---------------------------------------------------------------------------
# ====================== 主窗口 ======================
# ---------------------------------------------------------------------------
class ServerTestWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.selected_ramp_up = 1
        self.selected_duration = 3
        self.selected_concurrent_users = 1
        self.selected_threads_num = 1
        self.selected_operate_environment = ENVIRONMENT_TEST
        self.selected_operate_interval = 1
        self.selected_execute_times = 1
        self.test_data_table = None
        self.script_path = None
        self.stop_test = False
        self.pause_test = False

        self.node_to_index = {}
        self.executed_cases_count = 0
        self.collector_thread = None
        self.runner_thread = None

        base = os.path.dirname(os.path.abspath(__file__))
        ui_path = os.path.join(base, "server_test.ui")
        if not os.path.exists(ui_path):
            ui_path = os.path.join(base, "ui", "server_test.ui")
        loadUi(ui_path, self)

        flags = self.windowFlags()
        flags &= ~Qt.WindowMaximizeButtonHint
        flags |= Qt.WindowSystemMenuHint | Qt.WindowTitleHint
        self.setWindowFlags(flags)

        self._init_ui_widgets()
        self._bind_all_events()
        self._init_manager()

    # ------------------------------------------------------------------
    # UI 初始化
    # ------------------------------------------------------------------
    def _init_ui_widgets(self):
        self.setMenuItemStyle(self.menu_file)
        self.setMenuItemStyle(self.menu_config)
        self.setMenuItemStyle(self.menu_theme)
        self.setMenuItemStyle(self.menu_help)

        self.log_text_edit.setReadOnly(True)
        self.log_text_edit.setFont(QFont("Consolas", 10))

        self.operate_environment_combo.clear()
        self.operate_environment_combo.addItems(DEFAULT_OPERATE_ENVIRONMENT_OPTIONS)

        self.execute_times_combo.clear()
        self.execute_times_combo.addItems(DEFAULT_EXECUTE_TIMES_OPTIONS)
        self.operate_interval_combo.clear()
        self.operate_interval_combo.addItems(DEFAULT_OPERATE_INTERVAL_OPTIONS)
        self.threads_num_combo.clear()
        self.threads_num_combo.addItems(DEFAULT_NUM_OF_THREADS_OPTIONS)

        self.concurrent_users_combo.clear()
        self.concurrent_users_combo.addItems(DEFAULT_NUM_OF_CONCURRENT_USERS_OPTIONS)
        self.duration_combo.clear()
        self.duration_combo.addItems(DEFAULT_DURATION_OPTIONS)
        self.ramp_up_combo.clear()
        self.ramp_up_combo.addItems(DEFAULT_RAMP_UP_OPTIONS)

        self.test_data_table = QTableWidget(self.test_data_group)
        self.test_data_table.setGeometry(10, 30, 800, 580)
        self.test_data_table.setColumnCount(3)
        self.test_data_table.setHorizontalHeaderLabels(TABLE_HEADERS)
        self.test_data_table.setRowCount(0)
        self.test_data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.test_data_table.setItemDelegate(self.CenterAlignDelegate())
        self.test_data_table.verticalHeader().setVisible(False)

        self.test_progress_bar.setRange(0, 100)
        self.test_progress_bar.setValue(0)
        self.test_progress_bar.setStyleSheet("QProgressBar { color: #000; font-weight: bold; }")

        self.total_case_value.setText("0条")
        self.success_case_value.setText("0条")
        self.fail_case_value.setText("0条")
        self.skip_case_value.setText("0条")

    def _bind_all_events(self):
        self.operate_environment_combo.currentTextChanged.connect(self.on_operate_environment_selected)

        self.execute_times_combo.currentTextChanged.connect(self.on_execute_times_selected)
        self.operate_interval_combo.currentTextChanged.connect(self.on_operate_interval_selected)
        self.threads_num_combo.currentTextChanged.connect(self.on_threads_num_selected)

        self.concurrent_users_combo.currentTextChanged.connect(self.on_concurrent_users_selected)
        self.duration_combo.currentTextChanged.connect(self.on_duration_selected)
        self.ramp_up_combo.currentTextChanged.connect(self.on_ramp_up_selected)

        self.log_copy_btn.clicked.connect(self.on_copy_log)
        self.log_clear_btn.clicked.connect(self.on_clear_log)
        self.log_save_btn.clicked.connect(self.on_save_log)
        self.start_test_btn.clicked.connect(self.on_start_test)
        self.pause_test_btn.clicked.connect(self.on_pause_test)
        self.stop_test_btn.clicked.connect(self.on_stop_test)
        self.log_level_btn.clicked.connect(self.on_log_level_clicked)

        self.action_loadScript.triggered.connect(self.on_load_script)
        self.action_exportReport.triggered.connect(self.on_export_report)
        self.action_exit.triggered.connect(self.close)
        self.action_viewConfigFile.triggered.connect(self.on_view_config)
        self.action_editConfigFile.triggered.connect(self.on_edit_config)
        self.action_blackTheme.triggered.connect(self.on_theme_black)
        self.action_greenTheme.triggered.connect(self.on_theme_green)
        self.action_defaultTheme.triggered.connect(self.on_theme_default)
        self.action_about.triggered.connect(self.on_about)

    def _init_manager(self):
        self.selogger = ServerHandLogger(self.log_text_edit)

    # ------------------------------------------------------------------
    # 脚本加载
    # ------------------------------------------------------------------
    def on_load_script(self):
        self.selogger.log(f"正在选择测试脚本...")
        base_dir = os.path.dirname(os.path.abspath(__file__))
        scripts_dir = os.path.join(base_dir, "scripts")
        if not os.path.exists(scripts_dir):
            os.makedirs(scripts_dir)

        file_path, _ = QFileDialog.getOpenFileName(self, "选择测试脚本", scripts_dir, "Python files (*.py)")
        if not file_path:
            return

        self.script_path = file_path
        self.test_data_table.setRowCount(0)
        self.node_to_index.clear()

        self.selogger.log(f"正在解析脚本：{file_path}")
        self.start_test_btn.setEnabled(False)

        self.collector_thread = TestCollectorThread(file_path)
        self.collector_thread.sig_collected.connect(self.render_cases_to_table)
        self.collector_thread.sig_error.connect(lambda e: self.selogger.log(f"解析失败：{e}", level="ERROR"))
        self.collector_thread.start()

    def render_cases_to_table(self, cases):
        cnt = len(cases)
        if cnt == 0:
            self.selogger.log("未发现测试用例！", level="ERROR")
            self.start_test_btn.setEnabled(True)
            return

        for item in cases:
            row = item["index"]
            nodeid = item["nodeid"]
            self.node_to_index[nodeid] = row

            self.test_data_table.insertRow(row)
            id_item = QTableWidgetItem(f'用例{row + 1}')
            id_item.setTextAlignment(Qt.AlignCenter)
            self.test_data_table.setItem(row, 0, id_item)

            name_item = QTableWidgetItem(item["name"])
            name_item.setToolTip(item["nodeid"])
            self.test_data_table.setItem(row, 1, name_item)

            status_item = QTableWidgetItem("待执行")
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QBrush(QColor("#909399")))
            self.test_data_table.setItem(row, 2, status_item)

        self.total_case_value.setText(f"{cnt}条")
        self.success_case_value.setText("0条")
        self.fail_case_value.setText("0条")
        self.skip_case_value.setText("0条")
        self.test_progress_bar.setValue(0)
        self.start_test_btn.setEnabled(True)
        self.selogger.log(f"脚本加载完成，共 {cnt} 条用例")

    # ------------------------------------------------------------------
    # 开始 / 暂停 / 停止 / 状态更新
    # ------------------------------------------------------------------
    def on_start_test(self):
        if not self.script_path or self.test_data_table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "请先加载测试脚本！")
            return

        self.stop_test = False
        self.pause_test = False
        OperateSharedData.write_control(stop_test=False, pause_test=False)
        self.reset_all_case_status()

        self.start_test_btn.setEnabled(False)
        self.pause_test_btn.setEnabled(True)
        self.stop_test_btn.setEnabled(True)
        self.pause_test_btn.setText("暂停测试")

        self.success_case_value.setText("0条")
        self.fail_case_value.setText("0条")
        self.skip_case_value.setText("0条")
        self.test_progress_bar.setValue(0)
        self.executed_cases_count = 0

        self.selogger.log("=" * 50)
        self.selogger.log("开始执行自动化测试...")

        self.runner_thread = TestRunnerThread(self.script_path, self.node_to_index)
        self.runner_thread.sig_log.connect(self.log_from_engine)
        self.runner_thread.sig_status.connect(self.update_case_status)
        self.runner_thread.sig_finished.connect(self.on_test_all_finished)
        self.runner_thread.start()

    def reset_all_case_status(self):
        row_count = self.test_data_table.rowCount()
        for row in range(row_count):
            item = self.test_data_table.item(row, 2)
            if item:
                item.setText("待测试")
                item.setForeground(QBrush(Qt.black))

    def on_pause_test(self):
        if not self.runner_thread or not self.runner_thread.isRunning():
            self.status_bar.showMessage('当前无测试任务运行')
            self.selogger.log("暂无运行中的测试任务")
            return

        self.pause_test = not self.pause_test
        self.runner_thread.set_pause(self.pause_test)
        OperateSharedData.write_control(stop_test=False, pause_test=self.pause_test)

        if self.pause_test:
            self.pause_test_btn.setText("继续测试")
            self.status_bar.showMessage("测试已暂停 → 点击继续")
            self.selogger.log("⏸️ 测试已暂停")
        else:
            self.pause_test_btn.setText("暂停测试")
            self.status_bar.showMessage("测试已恢复执行")
            self.selogger.log("▶️ 测试已恢复")

    def on_stop_test(self):
        if not self.runner_thread or not self.runner_thread.isRunning():
            self.status_bar.showMessage('当前无测试任务运行')
            self.selogger.log("暂无运行中的测试任务")
            return

        self.selogger.log("🛑 正在停止测试...")
        self.stop_test = True
        self.pause_test = True

        self.runner_thread.set_pause(True)
        self.runner_thread.set_stop(True)
        OperateSharedData.write_control(stop_test=True, pause_test=True)

        self.pause_test_btn.setText("暂停测试")
        self.pause_test_btn.setEnabled(False)
        self.stop_test_btn.setEnabled(False)
        self.start_test_btn.setEnabled(True)

        self.status_bar.showMessage("测试已停止")
        self.selogger.log("✅ 测试已完全停止")

    def update_case_status(self, row, status):
        item = self.test_data_table.item(row, 2)
        total = self.test_data_table.rowCount()

        if status == "RUNNING":
            item.setText("执行中")
            item.setForeground(QBrush(QColor("#E6A23C")))
        elif status == "PASS":
            item.setText("通过")
            item.setForeground(QBrush(QColor("#67C23A")))
            v = self.success_case_value.text().replace("条", "")
            self.success_case_value.setText(f"{int(v) + 1}条")
        elif status == "FAIL":
            item.setText("失败")
            item.setForeground(QBrush(QColor("#F56C6C")))
            v = self.fail_case_value.text().replace("条", "")
            self.fail_case_value.setText(f"{int(v) + 1}条")
        elif status == "SKIP":
            item.setText("跳过")
            item.setForeground(QBrush(QColor("#E6A23C")))
            v = self.skip_case_value.text().replace("条", "")
            self.skip_case_value.setText(f"{int(v) + 1}条")

        if status in ["PASS", "FAIL", "SKIP"] and not self.pause_test and not self.stop_test:
            self.executed_cases_count += 1
            progress = int((self.executed_cases_count / total) * 100)
            self.test_progress_bar.setValue(progress)

    def log_from_engine(self, log):
        msg = log.get("msg")
        if msg:
            print(msg)

    def on_test_all_finished(self):
        self.test_progress_bar.setValue(100)
        self.start_test_btn.setEnabled(True)
        self.pause_test_btn.setEnabled(False)
        self.stop_test_btn.setEnabled(False)
        self.selogger.log("✅ 所有用例执行完毕！")

    # ------------------------------------------------------------------
    # 工具函数
    # ------------------------------------------------------------------
    def on_copy_log(self):
        self.log_text_edit.selectAll()
        self.log_text_edit.copy()

    def on_clear_log(self):
        self.log_text_edit.clear()

    def on_save_log(self):
        self.selogger.log("开始保存日志文件")
        filename = f"服务器测试日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        fn, _ = QFileDialog.getSaveFileName(self, "保存日志", filename, "文本文件 (*.txt)")
        if fn:
            try:
                with open(fn, "w", encoding="utf-8") as f:
                    f.write(self.log_text_edit.toPlainText())
                self.selogger.log(f"日志保存成功：{os.path.basename(fn)}")
            except Exception as e:
                self.selogger.log(f"日志保存失败：{str(e)}", level="ERROR")

    def on_log_level_clicked(self):
        menu = QMenu(self)
        self.setMenuItemStyle(menu)
        log_levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']
        current_level = getattr(self.selogger, 'log_level', 'INFO')
        for level in log_levels:
            action = QAction(level, self)
            if level == current_level:
                action.setCheckable(True)
                action.setChecked(True)
            action.triggered.connect(lambda checked, l=level: self.on_log_level_selected(l))
            menu.addAction(action)
        pos = self.log_level_btn.mapToGlobal(QPoint(0, self.log_level_btn.height()))
        menu.exec_(pos)

    def on_log_level_selected(self, level):
        try:
            self.selogger.set_log_level(level)
            self.selogger.log(f"日志级别已设置为：{level}")
        except Exception as e:
            self.selogger.log(f"设置日志级别失败：{str(e)}")

    def on_export_report(self):
        """导出测试报告为Excel文件 —— 修复版：读取真实测试用例数据"""
        row_count = self.test_data_table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "提示", "请先加载脚本并执行测试，再导出报告！")
            self.selogger.log("导出失败：未加载测试脚本或无测试数据")
            return

        self.selogger.log("开始导出测试报告...")

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16)

        wb = Workbook()
        ws = wb.active
        ws.title = "灵巧手测试报告"

        # ===================== 标题 =====================
        ws.merge_cells('A1:C1')
        ws['A1'] = "APP自动化测试报告"
        ws['A1'].font = title_font
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 30

        # ===================== 基础信息 =====================
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A3'] = "测试时间"
        ws['B3'] = now
        ws['A4'] = "测试脚本"
        ws['B4'] = os.path.basename(self.script_path) if self.script_path else "未选择"
        ws['A5'] = "总用例数"
        ws['B5'] = row_count

        # 统计结果
        success = int(self.success_case_value.text().replace("条", ""))
        fail = int(self.fail_case_value.text().replace("条", ""))
        skip = int(self.skip_case_value.text().replace("条", ""))
        ws['A6'] = "通过"
        ws['B6'] = success
        ws['A7'] = "失败"
        ws['B7'] = fail
        ws['A8'] = "跳过"
        ws['B8'] = skip

        # ===================== 表头 =====================
        headers = ["用例编号", "用例名称", "执行状态"]
        ws.append([""])  # 空行
        ws.append(headers)
        header_row = 10
        for col, value in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=value)
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # ===================== 写入真实测试用例数据 =====================
        for row in range(row_count):
            case_id = self.test_data_table.item(row, 0).text() if self.test_data_table.item(row, 0) else ""
            case_name = self.test_data_table.item(row, 1).text() if self.test_data_table.item(row, 1) else ""
            case_status = self.test_data_table.item(row, 2).text() if self.test_data_table.item(row, 2) else ""

            ws.append([case_id, case_name, case_status])

            # 给状态上色（Excel里也显示颜色）
            data_row = ws.max_row
            status_cell = ws.cell(row=data_row, column=3)
            if "通过" in case_status:
                status_cell.font = Font(color="67C23A", bold=True)
            elif "失败" in case_status:
                status_cell.font = Font(color="F56C6C", bold=True)
            elif "执行中" in case_status or "跳过" in case_status:
                status_cell.font = Font(color="E6A23C", bold=True)

        # ===================== 列宽/样式 =====================
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15

        for r in range(header_row, ws.max_row + 1):
            for c in range(1, 4):
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).alignment = center

        # ===================== 保存文件 =====================
        os.makedirs("report", exist_ok=True)
        time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        if self.script_path:
            script_name = os.path.splitext(os.path.basename(self.script_path))[0]
            filename = f"测试报告_{script_name}_{time_str}.xlsx"
        else:
            filename = f"测试报告_{time_str}.xlsx"

        save_path = os.path.join("report", filename)

        try:
            wb.save(save_path)
            self.selogger.log(f"报告导出成功：{save_path}")
            QMessageBox.information(self, "导出成功", f"测试报告已保存至：\n{save_path}")
        except Exception as e:
            self.selogger.log(f"报告导出失败：{str(e)}", level="ERROR")
            QMessageBox.critical(self, "导出失败", f"错误：{str(e)}")

    def on_view_config(self):
        config_path = ServerManager.get_configfile_path()
        if not os.path.exists(config_path):
            QMessageBox.warning(self, "警告", f"配置文件不存在：{config_path}")
            self.selogger.log(f"配置文件不存在：{config_path}")
            return
        try:
            with open(config_path, 'r', encoding="UTF-8") as f:
                content = f.read()
            dialog = QDialog(self)
            dialog.setWindowTitle("查看配置文件")
            dialog.setMinimumSize(600, 800)
            layout = QVBoxLayout()
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setFont(QFont("Consolas", 10))
            text_edit.setPlainText(content)
            text_edit.setStyleSheet("""
                QTextEdit { background-color: white; color: #333; border: 2px solid #d1d5db; border-radius: 6px; padding: 8px; font-size: 14px; }
            """)
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(text_edit)
            layout.addWidget(button_box)
            dialog.setLayout(layout)
            dialog.exec_()
            self.selogger.log(f"成功查看配置文件：{config_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取配置文件失败：{str(e)}")
            self.selogger.log(f"读取配置文件失败：{str(e)}")

    def on_edit_config(self):
        config_path = ServerManager.get_configfile_path()
        if not os.path.exists(config_path):
            QMessageBox.warning(self, "警告", f"配置文件不存在：{config_path}")
            self.selogger.log(f"配置文件不存在：{config_path}")
            return
        try:
            with open(config_path, 'r', encoding="UTF-8") as f:
                content = f.read()
            dialog = QDialog(self)
            dialog.setWindowTitle("修改配置文件")
            dialog.setMinimumSize(600, 800)
            layout = QVBoxLayout()
            hint_label = QLabel("提示：修改后需要重启应用程序才能生效")
            hint_label.setStyleSheet("color: #f59e0b; font-weight: bold; padding: 5px;")
            layout.addWidget(hint_label)
            text_edit = QTextEdit()
            text_edit.setFont(QFont("Consolas", 10))
            text_edit.setPlainText(content)
            text_edit.setStyleSheet("""
                QTextEdit { background-color: white; color: #333; border: 2px solid #d1d5db; border-radius: 6px; padding: 8px; font-size: 14px; }
            """)
            button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
            save_btn = button_box.button(QDialogButtonBox.Save)
            save_btn.setText("保存")
            save_btn.clicked.connect(lambda: self._save_config_and_restart(config_path, text_edit.toPlainText(), dialog))
            cancel_btn = button_box.button(QDialogButtonBox.Cancel)
            cancel_btn.setText("取消")
            cancel_btn.clicked.connect(dialog.reject)
            layout.addWidget(text_edit)
            layout.addWidget(button_box)
            dialog.setLayout(layout)
            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"编辑配置文件失败：{str(e)}")
            self.selogger.log(f"编辑配置文件失败：{str(e)}")

    def _save_config_and_restart(self, config_path, content, dialog):
        try:
            with open(config_path, 'w', encoding="UTF-8") as f:
                f.write(content)
            self.selogger.log(f"配置文件已保存：{config_path}")
            QMessageBox.information(self, "成功", "配置文件已保存，应用程序将自动重启...")
            dialog.accept()
            script_path = os.path.abspath(sys.argv[0])
            subprocess.Popen([sys.executable, script_path], start_new_session=True,
                             creationflags=subprocess.DETACHED_PROCESS if os.name == 'nt' else 0)
            time.sleep(0.5)
            QApplication.quit()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")
            self.selogger.log(f"保存配置文件失败：{str(e)}")

    def on_operate_environment_selected(self, text):
        try:
            environment_value = text
            self.selogger.log(f"已选择 测试环境：{text}")
            self.selected_operate_environment = environment_value
            OperateSharedData.write_environment_params(operate_environment=self.selected_operate_environment)
        except Exception as e:
            self.selogger.log(f"选择测试环境异常：{e}")

    def on_execute_times_selected(self, text):
        try:
            times_value = int(text.replace("次", ""))
            self.selogger.log(f"已选择 执行次数：{text}")
            self.selected_execute_times = times_value
            OperateSharedData.write_fun_params(execute_times=self.selected_execute_times)
        except Exception as e:
            self.selogger.log(f"选择执行次数异常：{e}")

    def on_operate_interval_selected(self, text):
        try:
            interval_value = int(text.replace("秒", ""))
            self.selogger.log(f"已选择 操作间隔：{text}")
            self.selected_operate_interval = interval_value
            OperateSharedData.write_fun_params(operate_interval=self.selected_operate_interval)
        except Exception as e:
            self.selogger.log(f"选择间隔异常：{e}")

    def on_threads_num_selected(self, text):
        try:
            threads_num_value = int(text)
            self.selogger.log(f"已选择 线程数：{text}")
            self.selected_threads_num = threads_num_value
            OperateSharedData.write_fun_params(threads_num=self.selected_threads_num)
        except Exception as e:
            self.selogger.log(f"选择线程数异常：{e}")

    def on_concurrent_users_selected(self, text):
        try:
            concurrent_users_value = int(text)
            self.selogger.log(f"已选择 并发用户数：{text}")
            self.selected_concurrent_users = concurrent_users_value
            OperateSharedData.write_perf_params(concurrent_user_nums=self.selected_concurrent_users)
        except Exception as e:
            self.selogger.log(f"选择并发用户数异常：{e}")

    def on_duration_selected(self, text):
        try:
            duration_value = int(text.replace("分钟", ""))
            self.selogger.log(f"已选择 运行时间：{text}")
            self.selected_duration = duration_value
            OperateSharedData.write_perf_params(duration=self.selected_duration)
        except Exception as e:
            self.selogger.log(f"选择运行时间异常：{e}")

    def on_ramp_up_selected(self, text):
        try:
            ramp_up_value = int(text.replace("秒", ""))
            self.selogger.log(f"已选择 爬坡时间：{text}")
            self.selected_ramp_up = ramp_up_value
            OperateSharedData.write_perf_params(ramp_up=self.selected_ramp_up)
        except Exception as e:
            self.selogger.log(f"选择爬坡时间异常：{e}")

    def on_theme_black(self):
        self.selogger.log("切换至黑色主题")
        apply_black_style(self)

    def on_theme_green(self):
        self.selogger.log("切换至绿色主题")
        apply_green_style(self)

    def on_theme_default(self):
        self.selogger.log("恢复默认主题")
        apply_default_style(self)

    def closeEvent(self, event):
        try:
            OperateSharedData.delete_shared_data_file()
        except:
            pass
        event.accept()

    def setMenuItemStyle(self, menu: QMenu):
        menu.setStyleSheet("""
            QMenu::item { height: 20px; width: 200px; margin: 0px; font-size: 16px; }
        """)

    class CenterAlignDelegate(QStyledItemDelegate):
        def initStyleOption(self, option, index):
            super().initStyleOption(option, index)
            option.displayAlignment = Qt.AlignCenter

    def on_about(self):
        QMessageBox.about(self, "关于", "灵巧手自动化测试工具 v1.0\n基于PyQt5开发")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = ServerTestWindow()
    win.setWindowTitle("服务器自动化测试界面")
    win.show()
    sys.exit(app.exec_())